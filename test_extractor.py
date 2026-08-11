import pytest
import pandas as pd
import numpy as np
import datetime
from io import BytesIO
import openpyxl
import os

from extractor import (
    load_data,
    clean_date,
    clean_time,
    extract_course_code_and_title,
    detect_header_row,
    detect_timetable_sheets,
    score_header_row,
    create_excel_download,
    create_pdf_download
)

POSSIBLE_TIMETABLE_PATHS = [
    "FINAL Bachelor's Virtual Exam Timetable - Aug 2026.(1).xlsx",
    "/home/paul/Downloads/FINAL Bachelor's Virtual Exam Timetable - Aug 2026.(1).xlsx",
    "/home/paul/Downloads/FINAL Bachelor's Virtual Exam Timetable - Aug 2026..xlsx"
]

def get_real_timetable_path():
    for p in POSSIBLE_TIMETABLE_PATHS:
        if os.path.exists(p):
            return p
    return None

# ==========================================
# 1. Test Real Timetable Regression
# ==========================================

def test_real_timetable_loading():
    real_path = get_real_timetable_path()
    if not real_path:
        pytest.skip("Sample timetable spreadsheet file not found in workspace or Downloads.")

    with open(real_path, "rb") as f:
        file_bytes = f.read()

    df_clean, selected_sheet, header_row_idx, all_sheets = load_data(file_bytes)

    assert selected_sheet == " BACHELOR'S FINAL EXAM TT"
    assert header_row_idx == 5
    assert len(df_clean) >= 450
    assert "COURSE_CODE" in df_clean.columns
    assert "COURSE_TITLE" in df_clean.columns
    assert "EXAMS_DATE" in df_clean.columns
    assert "SESSION_TIME" in df_clean.columns

    codes = set(df_clean["COURSE_CODE"].unique())
    assert "BAF3111" in codes
    assert "BUCU007" in codes
    assert "BIT2103" in codes

    selected = ["BAF3111", "BUCU007"]
    filtered = df_clean[df_clean["COURSE_CODE"].isin(selected)]
    assert len(filtered) == 2


# ==========================================
# 2. Test Date & Time Cleaning Edge Cases
# ==========================================

def test_clean_date():
    assert clean_date(None) is None
    assert clean_date(np.nan) is None
    assert clean_date("") is None

    dt = datetime.datetime(2026, 8, 10, 0, 0)
    assert clean_date(dt) == datetime.date(2026, 8, 10)
    d = datetime.date(2026, 8, 10)
    assert clean_date(d) == d

    assert clean_date("2026-08-10 00:00:00") == datetime.date(2026, 8, 10)
    assert clean_date("MON 12/08/2026") == datetime.date(2026, 8, 12)
    assert clean_date("12th August 2026") == datetime.date(2026, 8, 12)
    assert clean_date("17th - 22nd Aug") == "17th - 22nd Aug"


def test_clean_time():
    assert clean_time(None) == ""
    assert clean_time(np.nan) == ""

    assert clean_time(datetime.time(8, 0)) == "08:00"
    assert clean_time(datetime.datetime(2026, 8, 10, 14, 30)) == "14:30"

    assert clean_time("08:00 AM") == "08:00 AM"
    assert clean_time("8.00 AM") == "8:00 AM"
    assert clean_time("8:00AM-10:00AM") == "8:00 AM - 10:00 AM"
    assert clean_time("ONLINE") == "ONLINE"
    assert clean_time("Morning") == "Morning"


# ==========================================
# 3. Test Course Code and Title Extraction
# ==========================================

def test_extract_course_code_and_title():
    code, title = extract_course_code_and_title("BAF 3111", "Property Insurance")
    assert code == "BAF3111"
    assert title == "Property Insurance"

    code, title = extract_course_code_and_title("CSC 401 - Database Systems")
    assert code == "CSC401"
    assert title == "Database Systems"

    code, title = extract_course_code_and_title("CSC401 Database Systems")
    assert code == "CSC401"
    assert title == "Database Systems"

    code, title = extract_course_code_and_title("CSC 401: Database Systems")
    assert code == "CSC401"
    assert title == "Database Systems"

    code, title = extract_course_code_and_title("BUCU007")
    assert code == "BUCU007"


# ==========================================
# 4. Test Alternative Column Names & Offset Header
# ==========================================

def test_alternative_column_names_and_offset_header():
    data = []
    for i in range(12):
        data.append(["Intro line " + str(i), None, None, None, None])
    data.append(["Serial No", "Exam Date", "Session", "Unit Code", "Unit Name"])
    data.append([1, "2026-09-01", "09:00 AM", "MATH 101", "Calculus I"])
    data.append([2, "2026-09-01", "01:00 PM", "ENG 102", "Communication"])
    data.append([3, "2026-09-02", "09:00 AM", "CS 103", "Intro to Programming"])

    buf = BytesIO()
    df_build = pd.DataFrame(data)
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_build.to_excel(writer, index=False, header=False, sheet_name="Timetable")

    buf.seek(0)
    df_clean, selected_sheet, header_row_idx, _ = load_data(buf.read())

    assert selected_sheet == "Timetable"
    assert header_row_idx == 12
    assert len(df_clean) == 3
    assert set(df_clean["COURSE_CODE"]) == {"MATH101", "ENG102", "CS103"}
    assert set(df_clean["COURSE_TITLE"]) == {"Calculus I", "Communication", "Intro to Programming"}


# ==========================================
# 5. Test Merged Date & Time Cell Forward Filling
# ==========================================

def test_merged_date_and_time_forward_fill():
    data = [
        ["Code", "Date", "Time", "Venue"],
        ["CS 201", "2026-10-01", "08:00 AM", "Lab 1"],
        ["CS 202", None, None, "Lab 2"],
        ["CS 203", None, "02:00 PM", "Lab 1"],
    ]
    buf = BytesIO()
    pd.DataFrame(data).to_excel(buf, index=False, header=False, engine="openpyxl")
    buf.seek(0)

    df_clean, _, _, _ = load_data(buf.read())
    assert len(df_clean) == 3
    assert df_clean.iloc[0]["EXAMS_DATE"] == datetime.date(2026, 10, 1)
    assert df_clean.iloc[1]["EXAMS_DATE"] == datetime.date(2026, 10, 1)
    assert df_clean.iloc[1]["SESSION_TIME"] == "08:00 AM"
    assert df_clean.iloc[2]["SESSION_TIME"] == "02:00 PM"


# ==========================================
# 6. Test Multiple Sheets Selection
# ==========================================

def test_multiple_sheets_detection():
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame([["Welcome to the Exam Timetable System"], ["Author: Registrar"]]).to_excel(
            writer, index=False, header=False, sheet_name="Cover"
        )
        pd.DataFrame([
            ["S/No", "Date", "Time", "Unit Code", "Unit Name"],
            [1, "2026-08-10", "8:00 AM", "PHYS 101", "Physics I"]
        ]).to_excel(writer, index=False, header=False, sheet_name="Exam Schedule")

    buf.seek(0)
    file_bytes = buf.read()
    xl, sheet_scores, best_sheet = detect_timetable_sheets(file_bytes)

    assert best_sheet == "Exam Schedule"
    assert sheet_scores["Exam Schedule"] > sheet_scores["Cover"]


# ==========================================
# 7. Test Unrecognizable Timetable Error Handling
# ==========================================

def test_unrecognizable_timetable_error():
    buf = BytesIO()
    pd.DataFrame([
        ["Financial Summary", "2026"],
        ["Item A", 100],
        ["Item B", 200]
    ]).to_excel(buf, index=False, header=False, engine="openpyxl")

    buf.seek(0)
    with pytest.raises(ValueError) as exc_info:
        load_data(buf.read())

    assert "Could not recognize a valid exam timetable header" in str(exc_info.value)


# ==========================================
# 8. Test Excel Generation and Formatting
# ==========================================

def test_excel_export_formatting():
    df = pd.DataFrame([
        {
            "EXAMS_DATE": datetime.date(2026, 8, 10),
            "SESSION_TIME": "08:00 AM - 10:00 AM",
            "COURSE_CODE": "BAF3111",
            "COURSE_TITLE": "Property Insurance and Financial Risk Management",
            "VENUE": "Main Campus Hall 1"
        }
    ])

    excel_buf = create_excel_download(df)
    assert len(excel_buf.getvalue()) > 0

    wb = openpyxl.load_workbook(excel_buf)
    ws = wb.active

    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws.column_dimensions["A"].width == 15
    assert ws.column_dimensions["D"].width == 42
    assert ws["A1"].value == "Exam Date"
    assert ws["A2"].value == "10/08/2026"
    assert ws["D2"].value == "Property Insurance and Financial Risk Management"


# ==========================================
# 9. Test PDF Generation
# ==========================================

def test_pdf_export_generation():
    df = pd.DataFrame([
        {
            "EXAMS_DATE": datetime.date(2026, 8, 10),
            "SESSION_TIME": "08:00 AM - 10:00 AM",
            "COURSE_CODE": "BAF3111",
            "COURSE_TITLE": "Property Insurance and Financial Risk Management",
            "VENUE": "Main Campus Hall 1"
        }
    ])

    pdf_buf = create_pdf_download(df)
    pdf_bytes = pdf_buf.getvalue()
    assert len(pdf_bytes) > 0
    assert pdf_bytes.startswith(b"%PDF-")
