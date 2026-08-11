import streamlit as st
import pandas as pd
import numpy as np
import datetime
from io import BytesIO
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ==========================================
# 1. Normalization & Helper Functions
# ==========================================

def normalize_cell_value(val):
    """Safely converts any cell value (string, number, date, NaN, None) to a clean string."""
    if pd.isna(val) or val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date, datetime.time)):
        return str(val).strip()
    s = str(val).strip()
    if s.lower() in ("nan", "none", "null", "nat", "<na>"):
        return ""
    return s


def normalize_column_name(name):
    """Normalizes column header text for fuzzy keyword matching."""
    s = normalize_cell_value(name).upper()
    s = re.sub(r'[\s_\-/&,.():]+', ' ', s).strip()
    return s


KEYWORD_MAP = {
    'COURSE_CODE': [
        'UNIT CODE', 'COURSE CODE', 'MODULE CODE', 'SUBJECT CODE',
        'UNIT CODE NAME', 'COURSE CODE TITLE', 'UNIT CODE AND NAME',
        'COURSE CODE AND TITLE', 'UNIT CODE TITLE', 'COURSE_CODE', 'UNIT_CODE',
        'CODE'
    ],
    'COURSE_TITLE': [
        'UNIT NAME', 'COURSE TITLE', 'COURSE NAME', 'MODULE NAME',
        'SUBJECT NAME', 'UNIT TITLE', 'TITLE', 'DESCRIPTION', 'UNIT DESCRIPTION',
        'COURSE_TITLE', 'UNIT_NAME'
    ],
    'EXAMS_DATE': [
        'EXAMS DATE', 'EXAM DATE', 'DAY DATE', 'DAY AND DATE',
        'DAY DATE', 'EXAMS_DATE', 'EXAM_DATE', 'DATE', 'DAY'
    ],
    'SESSION_TIME': [
        'SESSION TIME', 'EXAM TIME', 'SESSION', 'TIME',
        'START TIME', 'END TIME', 'SLOT', 'DURATION', 'SESSION_TIME', 'EXAM_TIME'
    ],
    'VENUE': [
        'VENUE', 'ROOM', 'EXAM ROOM', 'LOCATION', 'HALL', 'LAB', 'CAMPUS', 'EXAM VENUE'
    ],
    'SNO': [
        'S NO', 'SNO', 'NO', 'SERIAL', 'SERIAL NO', 'ITEM', 'S/NO'
    ]
}


def score_header_row(row_values):
    """
    Scores a candidate header row based on how many semantic timetable categories it matches.
    Returns (score, category_mapping_dict) where category_mapping_dict is {col_index: category_name}.
    """
    category_matches = {}
    matched_cols = set()
    
    for col_idx, cell in enumerate(row_values):
        norm_cell = normalize_column_name(cell)
        if not norm_cell:
            continue
        
        best_cat = None
        for cat, keywords in KEYWORD_MAP.items():
            for kw in keywords:
                if norm_cell == kw or kw in norm_cell:
                    if cat == 'COURSE_CODE' and any(bad in norm_cell for bad in ['POSTAL', 'ZIP', 'COUNTRY', 'AREA']):
                        continue
                    best_cat = cat
                    break
            if best_cat:
                break
                
        if best_cat and best_cat not in matched_cols:
            matched_cols.add(best_cat)
            category_matches[col_idx] = best_cat

    score = 0
    if 'COURSE_CODE' in matched_cols:
        score += 3
    if 'EXAMS_DATE' in matched_cols:
        score += 2
    if 'SESSION_TIME' in matched_cols:
        score += 2
    if 'COURSE_TITLE' in matched_cols:
        score += 1
    if 'VENUE' in matched_cols:
        score += 1

    if 'COURSE_CODE' not in matched_cols:
        score = 0
        
    return score, category_matches


def detect_header_row(df_raw, max_rows=100):
    """
    Scans a dataframe without header up to max_rows to identify the header row.
    Returns (header_row_index, score, category_mapping_dict).
    """
    best_row_idx = None
    best_score = 0
    best_mapping = {}

    scan_limit = min(len(df_raw), max_rows)
    for idx in range(scan_limit):
        row_vals = df_raw.iloc[idx].values
        score, cat_map = score_header_row(row_vals)
        if score > best_score:
            best_score = score
            best_row_idx = idx
            best_mapping = cat_map

    return best_row_idx, best_score, best_mapping


# ==========================================
# 2. Field Cleaning Functions
# ==========================================

def clean_date(date_val):
    """
    Standardizes date values into consistent date objects or cleaned string representation.
    """
    if pd.isna(date_val) or date_val is None:
        return None
    if isinstance(date_val, datetime.datetime):
        return date_val.date()
    elif isinstance(date_val, datetime.date):
        return date_val

    d_str = str(date_val).strip()
    if not d_str or d_str.lower() in ("nan", "none", "null", "nat", "<na>"):
        return None

    m_iso = re.match(r'^(\d{4}-\d{2}-\d{2})', d_str)
    if m_iso:
        try:
            return pd.to_datetime(m_iso.group(1)).date()
        except Exception:
            pass

    if re.search(r'\d+(?:st|nd|rd|th)?\s*-\s*\d+(?:st|nd|rd|th)?', d_str, re.IGNORECASE) or ' TO ' in d_str.upper():
        return d_str

    cleaned = re.sub(
        r'^(MON|TUE|WED|THU|THUR|FRI|SAT|SUN|MONDAY|TUESDAY|WEDNESDAY|THURSDAY|FRIDAY|SATURDAY|SUNDAY)\s+',
        '', d_str, flags=re.IGNORECASE
    )
    cleaned = re.sub(r'(?<=\d)(st|nd|rd|th)', '', cleaned, flags=re.IGNORECASE)

    try:
        has_year = bool(re.search(r'\b(19|20)\d{2}\b', cleaned))
        dt = pd.to_datetime(cleaned, errors='coerce', dayfirst=True)
        if pd.notna(dt):
            if not has_year:
                dt = dt.replace(year=datetime.datetime.now().year)
            return dt.date()
    except Exception:
        pass

    return d_str


def clean_time(time_val):
    """
    Standardizes time or session values into clean display strings.
    """
    if pd.isna(time_val) or time_val is None:
        return ""
    if isinstance(time_val, datetime.time):
        return time_val.strftime("%H:%M")
    if isinstance(time_val, datetime.datetime):
        return time_val.strftime("%H:%M")
    t_str = str(time_val).strip()
    if not t_str or t_str.lower() in ("nan", "none", "null", "nat", "<na>"):
        return ""
    t_str = re.sub(r'(\d{1,2})\.(\d{2})', r'\1:\2', t_str)
    t_str = re.sub(r'(?<=[0-9A-Za-z])\s*-\s*(?=[0-9A-Za-z])', ' - ', t_str)
    t_str = re.sub(r'(\d{1,2}:\d{2})\s*([AP]M)', r'\1 \2', t_str, flags=re.IGNORECASE)
    return t_str


CODE_REGEX = r'([A-Za-z]{2,6}\s*[-_]?\s*\d{3,5}[A-Za-z]?)'

def extract_course_code_and_title(code_val, title_val=None):
    """
    Extracts canonical unit code and unit title.
    """
    code_str = str(code_val).strip() if pd.notna(code_val) and code_val is not None else ""
    title_str = str(title_val).strip() if pd.notna(title_val) and title_val is not None else ""
    
    if code_str.lower() in ("nan", "none", "null", "<na>", ""):
        return "", title_str

    match = re.search(CODE_REGEX, code_str)
    if match:
        extracted_code = match.group(1).upper()
        extracted_code = re.sub(r'\s+', '', extracted_code)
        
        if not title_str or title_str.lower() in ("nan", "none", "null", "<na>"):
            remains = code_str.replace(match.group(1), '').strip(' -:;,')
            if len(remains) > 2:
                title_str = remains
        return extracted_code, title_str
    else:
        if len(code_str) <= 20 and not re.search(r'\b(EXAM|SCHEDULE|PAGE|TOTAL|DATE|TIME|SNO|NO|CODE|TITLE|DAY|ROOM|VENUE)\b', code_str, re.IGNORECASE):
            return code_str.upper(), title_str
        return "", title_str


# ==========================================
# 3. Sheet Detection & Timetable Processing
# ==========================================

def detect_timetable_sheets(file_bytes, filename=""):
    """
    Inspects workbook sheets and scores candidates to find timetable sheets.
    """
    try:
        xl = pd.ExcelFile(BytesIO(file_bytes))
    except Exception as e:
        raise ValueError(f"Could not open spreadsheet file. Ensure it is a valid Excel/ODS file. Details: {e}")

    sheet_scores = {}
    best_sheet = None
    max_score = 0

    for sheet in xl.sheet_names:
        try:
            df_preview = pd.read_excel(xl, sheet_name=sheet, header=None, nrows=100)
            _, score, _ = detect_header_row(df_preview)
            sheet_scores[sheet] = score
            if score > max_score:
                max_score = score
                best_sheet = sheet
        except Exception:
            sheet_scores[sheet] = 0

    if best_sheet is None and xl.sheet_names:
        best_sheet = xl.sheet_names[0]

    return xl, sheet_scores, best_sheet


def clean_timetable(df_raw, header_row_idx, cat_map):
    """
    Processes raw dataframe slice from header_row_idx onwards into a clean exam timetable DataFrame.
    """
    new_cols = {}
    for col_idx in range(df_raw.shape[1]):
        if col_idx in cat_map:
            new_cols[col_idx] = cat_map[col_idx]
        else:
            new_cols[col_idx] = f"EXTRA_{col_idx}"

    data_df = df_raw.iloc[header_row_idx + 1:].rename(columns=new_cols).copy()

    for col in ['EXAMS_DATE', 'SESSION_TIME']:
        if col in data_df.columns:
            data_df[col] = data_df[col].ffill()

    processed_records = []
    header_words = {'COURSE_CODE', 'UNIT_CODE', 'CODE', 'COURSE CODE', 'UNIT CODE', 'COURSE_TITLE', 'UNIT_NAME', 'S/NO', 'EXAMS_DATE'}

    for _, row in data_df.iterrows():
        raw_code = row.get('COURSE_CODE', None)
        raw_title = row.get('COURSE_TITLE', None)

        code, title = extract_course_code_and_title(raw_code, raw_title)

        if not code or code.upper() in header_words:
            continue

        raw_date = row.get('EXAMS_DATE', None)
        raw_time = row.get('SESSION_TIME', None)
        c_date = clean_date(raw_date)
        c_time = clean_time(raw_time)

        venue_val = row.get('VENUE', None)
        venue = str(venue_val).strip() if pd.notna(venue_val) and venue_val is not None else ""
        if venue.lower() in ('nan', 'none', 'null', '<na>'):
            venue = ""

        processed_records.append({
            'EXAMS_DATE': c_date,
            'SESSION_TIME': c_time,
            'COURSE_CODE': code,
            'COURSE_TITLE': title,
            'VENUE': venue
        })

    df_clean = pd.DataFrame(processed_records)
    if not df_clean.empty:
        df_clean = df_clean.drop_duplicates(subset=['COURSE_CODE', 'EXAMS_DATE', 'SESSION_TIME', 'COURSE_TITLE'])
    return df_clean


def load_data(file_input, sheet_name=None):
    """
    Main loader function: reads spreadsheet, detects header, normalizes columns and cleans timetable records.
    """
    if isinstance(file_input, bytes):
        file_bytes = file_input
    else:
        file_input.seek(0)
        file_bytes = file_input.read()

    xl, sheet_scores, best_sheet = detect_timetable_sheets(file_bytes)
    selected_sheet = sheet_name if sheet_name and sheet_name in xl.sheet_names else best_sheet

    if not selected_sheet:
        raise ValueError("Workbook contains no readable sheets.")

    df_raw = pd.read_excel(xl, sheet_name=selected_sheet, header=None)
    header_row_idx, score, cat_map = detect_header_row(df_raw)

    if header_row_idx is None or score < 2:
        raise ValueError(
            f"Could not recognize a valid exam timetable header in sheet '{selected_sheet}'. "
            "Please check that the sheet has columns such as Unit Code, Course Code, Exam Date, or Time."
        )

    df_clean = clean_timetable(df_raw, header_row_idx, cat_map)
    return df_clean, selected_sheet, header_row_idx, xl.sheet_names


# ==========================================
# 4. Export & Download Format Helpers
# ==========================================

def create_excel_download(df):
    """
    Generates a beautifully formatted Excel workbook BytesIO buffer with auto-filters, custom column widths, and print setup.
    """
    cols = [c for c in ["EXAMS_DATE", "SESSION_TIME", "COURSE_CODE", "COURSE_TITLE", "VENUE"] if c in df.columns]
    df_export = df[cols].copy()

    if "EXAMS_DATE" in df_export.columns:
        df_export["EXAMS_DATE"] = df_export["EXAMS_DATE"].apply(
            lambda d: d.strftime("%d/%m/%Y") if isinstance(d, (datetime.date, datetime.datetime)) else (str(d) if pd.notna(d) and d is not None else "")
        )

    col_rename = {
        "EXAMS_DATE": "Exam Date",
        "SESSION_TIME": "Session Time",
        "COURSE_CODE": "Course Code",
        "COURSE_TITLE": "Course Title",
        "VENUE": "Venue"
    }
    df_export.rename(columns=col_rename, inplace=True)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Exam Timetable")

    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active

    col_widths = {
        "A": 15,  # Exam Date
        "B": 15,  # Session Time
        "C": 16,  # Course Code
        "D": 42,  # Course Title
        "E": 25   # Venue
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    body_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 24
        for cell in row:
            cell.font = body_font
            cell.border = thin_border
            if cell.column_letter in ["A", "B", "C", "E"]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"

    out_buffer = BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    return out_buffer


def create_pdf_download(df):
    """
    Generates a clean, landscape print-friendly PDF timetable BytesIO buffer.
    """
    cols = [c for c in ["EXAMS_DATE", "SESSION_TIME", "COURSE_CODE", "COURSE_TITLE", "VENUE"] if c in df.columns]
    df_export = df[cols].copy()

    if "EXAMS_DATE" in df_export.columns:
        df_export["EXAMS_DATE"] = df_export["EXAMS_DATE"].apply(
            lambda d: d.strftime("%d/%m/%Y") if isinstance(d, (datetime.date, datetime.datetime)) else (str(d) if pd.notna(d) and d is not None else "")
        )

    col_rename = {
        "EXAMS_DATE": "Exam Date",
        "SESSION_TIME": "Session Time",
        "COURSE_CODE": "Course Code",
        "COURSE_TITLE": "Course Title",
        "VENUE": "Venue"
    }
    df_export.rename(columns=col_rename, inplace=True)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#555555'),
        spaceAfter=12
    )
    header_cell_style = ParagraphStyle(
        'HeaderCell',
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        textColor=colors.white,
        alignment=1
    )
    body_cell_center = ParagraphStyle(
        'BodyCellCenter',
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#222222'),
        alignment=1
    )
    body_cell_left = ParagraphStyle(
        'BodyCellLeft',
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#222222'),
        alignment=0
    )

    headers = list(df_export.columns)
    table_data = [[Paragraph(h, header_cell_style) for h in headers]]

    for _, row in df_export.iterrows():
        row_data = []
        for col_name in headers:
            val_str = str(row.get(col_name, "") or "")
            if col_name == "Course Title":
                row_data.append(Paragraph(val_str, body_cell_left))
            else:
                row_data.append(Paragraph(val_str, body_cell_center))
        table_data.append(row_data)

    col_widths = [90, 110, 90, 270, 160]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements = [
        Paragraph("MKU Exam Timetable", title_style),
        Paragraph(f"Filtered Exam Schedule ({len(df_export)} units selected)", subtitle_style),
        Spacer(1, 6),
        t
    ]

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ==========================================
# 5. Streamlit User Interface
# ==========================================

def main():
    st.set_page_config(page_title="MKU Exam Extractor", page_icon="📅", layout="wide")
    st.title("MKU Exam Extractor")
    st.subheader("Extract and filter exam schedules for undergraduate and diploma courses.")
    st.write("Upload your exam timetable spreadsheet (.xlsx, .xls, .xlsm, .ods) and select your units.")

    uploaded_file = st.file_uploader(
        "Choose Excel/Spreadsheet file", 
        type=["xlsx", "xls", "xlsm", "ods"]
    )

    if uploaded_file is not None:
        try:
            file_bytes = uploaded_file.read()
            xl, sheet_scores, best_sheet = detect_timetable_sheets(file_bytes, uploaded_file.name)

            if len(xl.sheet_names) > 1:
                selected_sheet = st.selectbox(
                    "Select Timetable Worksheet",
                    options=xl.sheet_names,
                    index=xl.sheet_names.index(best_sheet) if best_sheet in xl.sheet_names else 0,
                    help="Auto-detected the sheet with the highest timetable confidence."
                )
            else:
                selected_sheet = best_sheet

            df_clean, sheet_used, header_row_idx, _ = load_data(file_bytes, sheet_name=selected_sheet)

            if df_clean.empty:
                st.error("No valid unit codes or exam records could be extracted from this worksheet.")
                return

            st.success(
                f"Successfully extracted {len(df_clean)} exam records from sheet **'{sheet_used}'** (Header detected at row {header_row_idx + 1})."
            )

            available_units = sorted(df_clean["COURSE_CODE"].dropna().unique().tolist())

            selected_units = st.multiselect(
                "Select your unit codes",
                options=available_units,
                help="Start typing or select unit codes to filter your personal timetable."
            )

            if selected_units:
                filtered = df_clean[df_clean["COURSE_CODE"].isin(selected_units)].copy()

                sort_cols = [c for c in ["EXAMS_DATE", "SESSION_TIME"] if c in filtered.columns]
                if sort_cols:
                    filtered = filtered.sort_values(by=sort_cols)

                cols_to_show = [c for c in ["EXAMS_DATE", "SESSION_TIME", "COURSE_CODE", "COURSE_TITLE", "VENUE"] if c in filtered.columns]
                
                preview_df = filtered[cols_to_show].copy()
                if "EXAMS_DATE" in preview_df.columns:
                    preview_df["EXAMS_DATE"] = preview_df["EXAMS_DATE"].apply(
                        lambda d: d.strftime("%d/%m/%Y") if isinstance(d, (datetime.date, datetime.datetime)) else (str(d) if pd.notna(d) and d is not None else "")
                    )

                st.subheader("Your Filtered Exam Timetable")
                st.dataframe(
                    preview_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "EXAMS_DATE": st.column_config.TextColumn("Exam Date"),
                        "SESSION_TIME": st.column_config.TextColumn("Session Time"),
                        "COURSE_CODE": st.column_config.TextColumn("Course Code"),
                        "COURSE_TITLE": st.column_config.TextColumn("Course Title"),
                        "VENUE": st.column_config.TextColumn("Venue")
                    }
                )

                st.subheader("Download your timetable:")
                excel_buffer = create_excel_download(filtered)
                pdf_buffer = create_pdf_download(filtered)

                dl_col1, dl_col2 = st.columns(2)
                with dl_col1:
                    st.download_button(
                        label="📄 Download as Excel",
                        data=excel_buffer,
                        file_name="my_exam_timetable.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                with dl_col2:
                    st.download_button(
                        label="🔴 Download as PDF",
                        data=pdf_buffer,
                        file_name="my_exam_timetable.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
            else:
                st.info("Please select at least one unit code above to view your timetable.")

        except Exception as e:
            st.error(f"Error processing timetable: {str(e)}")

if __name__ == "__main__":
    main()